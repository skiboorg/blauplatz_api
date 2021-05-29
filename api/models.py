import json

from django.db import models
from pytils.translit import slugify
from ckeditor_uploader.fields import RichTextUploadingField
from openpyxl import load_workbook
from django.db.models.signals import post_save,pre_save

class Richtungen(models.Model):
    name = models.CharField('Name', max_length=255, blank=False, null=True)
    name_slug = models.CharField(max_length=255, blank=True, null=True, db_index=True, editable=False)
    image = models.ImageField('Top image', blank=False, upload_to='richtungen/')
    icon = models.ImageField('Icon', blank=False, upload_to='richtungen/')
    text = RichTextUploadingField('Page text', blank=False, null=True)

    def save(self, *args, **kwargs):
        self.name_slug = slugify(self.name)
        super(Richtungen, self).save(*args, **kwargs)

    def __str__(self):
        return f'{self.name}'

class HilfreicheTabellen(models.Model):
    name = models.CharField('Name', max_length=255, blank=False, null=True)
    name_slug = models.CharField(max_length=255, blank=True, null=True, db_index=True, editable=False)
    file = models.FileField('Table', blank=False, upload_to='tabellen/')
    is_file_change = models.BooleanField('Is table changed',default=False)
    table = models.JSONField(blank=True,null=True)

    def save(self, *args, **kwargs):
        self.name_slug = slugify(self.name)
        print('save')

        super(HilfreicheTabellen, self).save(*args, **kwargs)

    def __str__(self):
        return f'{self.name}'

def table_post_save(sender, instance, created, **kwargs):
    if instance.is_file_change:
    # wb = load_workbook(filename='111.xlsx')

        wb = load_workbook(filename=instance.file)
        sheet = wb.active
        max_row = sheet.max_row
        data=[]
        max_column = sheet.max_column
        for row in range(1, max_row + 1):
            row_data={}
            col_data=[]
            for col in range(1, max_column + 1):
                cell_data = sheet.cell(row=row, column=col).value
                col_data.append(cell_data)
            row_data['row']=col_data
            data.append(row_data)

        instance.table = data
        instance.is_file_change = False
            # instance.table = json.dumps(data)
        instance.save()


post_save.connect(table_post_save, sender=HilfreicheTabellen)